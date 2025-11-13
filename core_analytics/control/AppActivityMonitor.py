"""
Simplified AppActivityMonitor controller for orchestration only.
"""
import datetime
from zoneinfo import ZoneInfo
import logging
import os

from core_analytics.config.settings import ConfigurationService
from core_analytics.model.repositories.azure_log_repository import AzureLogRepository
from core_analytics.services.analytics_service import AnalyticsService
from core_analytics.services.query_strategies.strategy_factory import QueryStrategyFactory
from core_analytics.view.factories.report_factory import ReportFactory
from core_analytics.view.factories.daily_monitor_factory import DailyMonitorFactory
from core_analytics.core.logging_config import LoggerSetup, CoreAnalyticsException
from core_analytics.services.email_service import EmailService
from core_analytics.services.cost_service import AzureCostService
from core_analytics.model.repositories.azure_blob_repository import AzureBlobRepository
from core_analytics.services.file_cleanup_service import FileCleanupService

from core_analytics.core.models import ProcessData

class AppActivityMonitor:
    """Main controller for orchestrating the analytics pipeline."""

    def __init__(self, days_range: int = 30):
        # Setup logging
        self.logger = LoggerSetup.setup_logger()
        
        # Calculate time range in JST, then convert to UTC for Azure API
        jst = ZoneInfo("Asia/Tokyo")
        end_jst = datetime.datetime.now(jst)
        start_jst = end_jst - datetime.timedelta(days=days_range)
        self.end_time = end_jst.astimezone(datetime.UTC)
        self.start_time = start_jst.astimezone(datetime.UTC)
        
        # Initialize dependencies
        self.config_service = ConfigurationService(days_range=days_range)
        self.log_repository = AzureLogRepository(self.config_service)
        self.strategy_factory = QueryStrategyFactory()
        self.analytics_service = AnalyticsService(
            self.log_repository, 
            self.config_service, 
            self.strategy_factory
        )
        self.report_factory = ReportFactory()
        self.daily_monitor_factory = DailyMonitorFactory()
        self.storage_service = AzureBlobRepository(self.config_service)
        self.file_cleanup_service = FileCleanupService(self.config_service)

        try:
            self.email_service = EmailService()
            self.email_enabled = True
        except Exception as e:
            self.logger.warning(f"Email service not available: {e}")
            self.email_service = None
            self.email_enabled = False

        try:
            self.cost_service = AzureCostService()
            self.cost_enabled = True
        except Exception as e:
            self.logger.warning(f"Cost service not available: {e}")
            self.cost_service = None
            self.cost_enabled = False
        
        self.logger.info("AppActivityMonitor initialized successfully")

    def run(self):
        """Main orchestration method - coordinates the entire analytics pipeline."""
        try:
            self.logger.info("Starting Core Analytics pipeline")
            self.logger.info(f"Processing data from {self.start_time} to {self.end_time}")

            report_mode = os.environ.get("REPORT_MODE", "daily_monitor")
            
            # Fetch and process analytics data
            processed_data : ProcessData = self.analytics_service.fetch_and_process_data(
                self.start_time, 
                self.end_time
            )
            
            # Setup output directory
            app_settings = self.config_service.get_app_settings()
            if report_mode == "standard":
                output_dir = f"{app_settings.output_base_dir}/{self.end_time.strftime('%Y%m%d')}"
                os.makedirs(output_dir, exist_ok=True)
            elif report_mode == "daily_monitor":
                output_dir = app_settings.output_base_dir
            
            generated_files = []

            if report_mode in ["standard"]:
                standard_files = self.report_factory.generate_all_reports(
                    processed_data, output_dir, self.end_time
                )
                generated_files.extend(standard_files)
                self.logger.info(f"Generated {len(standard_files)} standard reports")

            if report_mode in ["daily_monitor"]:
                self._download_existing_usage_report(output_dir)

                mtd_costs = None
                if self.cost_enabled:
                    mtd_costs = self.cost_service.get_apps_mtd_costs()
                    if mtd_costs:
                        self.logger.info(f"MTD costs: {mtd_costs}")
                
                daily_monitor_files = self.daily_monitor_factory.generate_daily_monitor_report(
                    processed_data, output_dir, self.end_time, mtd_costs
                )
                generated_files.extend(daily_monitor_files)
                self.logger.info(f"Generated {len(daily_monitor_files)} daily monitor reports")

                if self.email_enabled and daily_monitor_files:
                    date_str = self.end_time.strftime('%Y年%m月%d日')
                    email_sent = self.email_service.send_daily_monitor_report(
                        daily_monitor_files, date_str
                    )
                    if email_sent:
                        self.logger.info("Daily monitor report email sent successfully")
                    else:
                        self.logger.error("Failed to send daily monitor report email")

            #store the generated files to blob
            for file in generated_files:
                self.storage_service.upload_file(file, file)

            #cleanup the old files
            deleted_folders = self.file_cleanup_service.cleanup_old_output_directories(days_threshold=30)

            self.logger.info(f"Deleted {len(deleted_folders)} old output directories")
            self.logger.info(f"Analytics pipeline completed successfully")
            self.logger.info(f"Generated {len(generated_files)} report files")
            self.logger.info(f"Uploaded {len(generated_files)} report files to blob")
            
            # Print summary for user
            print(f"✅ Analytics pipeline completed successfully!")
            print(f"📊 Generated {len(generated_files)} reports in {output_dir}")
            print(f"📈 Processed data from {self.start_time.strftime('%Y-%m-%d')} to {self.end_time.strftime('%Y-%m-%d')}")
            print(f"Uploaded {len(generated_files)} report files to blob")
            
        except CoreAnalyticsException as e:
            self.logger.error(f"Analytics pipeline failed: {e}")
            print(f"❌ Pipeline failed: {e}")
            raise
            
        except Exception as e:
            self.logger.error(f"Unexpected error in analytics pipeline: {e}")
            print(f"❌ Unexpected error: {e}")
            raise CoreAnalyticsException(f"Pipeline failed with unexpected error: {e}")
    
    def _download_existing_usage_report(self, output_dir: str) -> None:
        """Ensure usage report exists in blob. If missing, seed blob with template, then download locally."""
        if os.environ.get("SKIP_USAGE_DOWNLOAD") == "1":
            return
        try:
            from pathlib import Path
            from openpyxl import load_workbook

            usage_report_path = Path("output/市場GAI打鍵/市場GAI使用状況.xlsx")
            usage_report_path.parent.mkdir(parents=True, exist_ok=True)
            blob_file_path = str(usage_report_path)

            try:
                self.storage_service.download_file(blob_file_path, str(usage_report_path))
                load_workbook(str(usage_report_path))
                self.logger.info(f"Downloaded existing usage report from blob: {usage_report_path}")
                return
            except Exception as download_error:
                self.logger.info(f"No existing usage report found in blob or download failed: {download_error}")

            template_path = Path("./config/report_template/市場GAI使用状況.xlsx")
            if not template_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_path}")

            self.storage_service.upload_file(str(template_path), blob_file_path)
            self.logger.info(f"Seeded usage report template to blob: {blob_file_path}")

            self.storage_service.download_file(blob_file_path, str(usage_report_path))
            load_workbook(str(usage_report_path))
            self.logger.info(f"Downloaded seeded usage report from blob to local: {usage_report_path}")

        except Exception as e:
            self.logger.error(f"Error during usage report download: {e}")


if __name__ == "__main__":
    app_activity_monitor = AppActivityMonitor()
    app_activity_monitor.run()
