import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference
from azure.monitor.query import LogsQueryResult
from datetime import datetime
from zoneinfo import ZoneInfo
import json
from pathlib import Path
from openpyxl.utils import get_column_letter
from datetime import timedelta
from typing import List

JST = ZoneInfo("Asia/Tokyo")

def _convert_timezone_aware_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    df_processed = df.copy()
    for column in df_processed.columns:
        if len(df_processed) > 0:
            sample_value = df_processed[column].iloc[0]
            if isinstance(sample_value, datetime) and sample_value.tzinfo is not None:
                df_processed[column] = (
                    df_processed[column]
                    .dt.tz_convert(JST)
                    .dt.tz_localize(None)
                )
    return df_processed

def _create_excel_from_LogsQueryResult(response:LogsQueryResult, filepath:str, sheet_name:str):
    df = pd.DataFrame(data=response.tables[0].rows, columns=response.tables[0].columns)
    df = _convert_timezone_aware_datetimes(df)
    df.to_excel(filepath, index=False, sheet_name=sheet_name)


def _read_user_count_excel_to_df(sheet_name: str, filepath :str)->tuple[pd.DataFrame, str]:
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    if "Department" in df.columns:
        column_name = "Department"
    return df, column_name

def _read_stroke_count_excel_to_df(sheet_name: str, filepath :str)->tuple[pd.DataFrame, str]:
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    return df


def _create_sheet_with_pie_chart(df:pd.DataFrame, filepath:str, column_name:str, sheet_name:str):
    wb = openpyxl.load_workbook(filepath)    

    dept_counts = df[column_name].value_counts().reset_index()

     # Excelにシートを作成
    ws2 = wb.create_sheet(sheet_name)

    # ヘッダーとデータ書き込み
    ws2.append(list(dept_counts.columns))
    for row in dept_counts.itertuples(index=False):
        ws2.append(list(row))

    # グラフ用データ範囲
    if ws2.max_row > 1:
        labels = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
        values = Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row)

        pie = PieChart()
        pie.add_data(values, titles_from_data=False)
        pie.set_categories(labels)
        pie.title = "Department割合"

        ws2.add_chart(pie, "E2")

    wb.save(filepath)

def _create_sheet_with_stroke_count(df:pd.DataFrame, filepath:str, sheet_name:str):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.create_sheet(sheet_name)

    #実行日から30日前までの連続日付を作成,土日を除く
    today = datetime.now(JST).date()
    start_date = today - timedelta(days=29)
    all_days = pd.date_range(start=start_date, end=today, freq="B")
    df_all = pd.DataFrame(all_days, columns=["stroke_date"])

    #マージ対象列のデータ型を統一
    df_all["stroke_date"] = df_all["stroke_date"].dt.date
    df["stroke_date"] = pd.to_datetime(df["stroke_date"]).dt.date

    #マージ
    df_all = df_all.merge(df, on="stroke_date", how="left")
    df_all["stroke_count"] = df_all["stroke_count"].fillna(0)
    ws.append(list(df_all.columns))
    for row in df_all.itertuples(index=False):
        ws.append(list(row))
    wb.save(filepath)


def generate_user_count_excel(response :LogsQueryResult, filepath :str):
    _create_excel_from_LogsQueryResult(response, filepath, "result")
    if "Department" in response.tables[0].columns:
        df, column_name = _read_user_count_excel_to_df("result", filepath)
        _create_sheet_with_pie_chart(df, filepath, column_name, "Department割合")

def generate_stroke_count_excel(response :LogsQueryResult, filepath :str):
    #LogsQueryResultをExcelに出力
    _create_excel_from_LogsQueryResult(response, filepath, "result")

    #Excelを読み込み、今回の打鍵数を取得する
    df = _read_stroke_count_excel_to_df("result", filepath)
    df["TimeGenerated"] = pd.to_datetime(df["TimeGenerated"]).dt.date
    df_stroke_count = df.groupby("TimeGenerated").size().reset_index(name="stroke_count").rename(columns={"TimeGenerated": "stroke_date"})
    
    _create_sheet_with_stroke_count(df_stroke_count, filepath, "打鍵数")


def add_row_to_stroke_count_excel(stroke_count:int, filepath:str, sheet_name:str):
    # ファイルの存在確認
    file_path = Path(f"{filepath}/stroke_count.xlsx")
    
    if file_path.exists():
        # ファイルが存在する場合は開く シートがあるかどうか確認
        wb = openpyxl.load_workbook(file_path)
        
    else:
        # ファイルが存在しない場合は新規作成
        wb = openpyxl.Workbook()
        # デフォルトシートを削除
        wb.remove(wb.active)

    sheet_exists = False
    # シート名の存在確認
    for sheet in wb.sheetnames:
        if sheet == sheet_name:
            sheet_exists = True
            break
    
    if sheet_exists:
        ws = wb[sheet_name]
    else:
        # シートが存在しない場合は新規作成
        ws = wb.create_sheet(sheet_name)
        ws.append(["stroke_date", "stroke_count"])
    
    # データを追加
    ws.append([datetime.now(JST).strftime("%Y-%m-%d"), stroke_count])
    # ファイル名を指定して保存
    wb.save(file_path)

def generate_stroke_count_summary_excel(datas:pd.DataFrame, filepath:str, sheet_name:str):
    # ファイルの存在確認
    file_path = Path(f"{filepath}/stroke_count.xlsx")
    
    if file_path.exists():
        # ファイルが存在する場合は開く シートがあるかどうか確認
        wb = openpyxl.load_workbook(file_path)
        
    else:
        # ファイルが存在しない場合は新規作成
        wb = openpyxl.Workbook()
        # デフォルトシートを削除
        wb.remove(wb.active)

    sheet_exists = False
    # シート名の存在確認
    for sheet in wb.sheetnames:
        if sheet == "Amount of stroke":
            sheet_exists = True
            break
    
    if sheet_exists:
        ws = wb["Amount of stroke"]
    else:
        # シートが存在しない場合は新規作成
        ws = wb.create_sheet("Amount of stroke")

        #実行日から30日前までの連続日付を作成,土日を除く
        today = datetime.now(JST).date()
        start_date = today - timedelta(days=29)
        all_days = pd.date_range(start=start_date, end=today, freq="B")
        all_days_array = all_days.strftime("%Y-%m-%d").tolist()
        all_days_array.insert(0, " ")

        ws.append(all_days_array)
    
    # データを追加
    data_array = datas["stroke_count"].to_list()
    data_array.insert(0, sheet_name)
    ws.append(data_array)
    # ファイル名を指定して保存
    wb.save(file_path)

def add_bar_graph_to_stroke_count_excel(folder_path:str):
    file_path = Path(f"{folder_path}/stroke_count.xlsx")
    wb = openpyxl.load_workbook(file_path)

    sheet_exists = False

    for sheet in wb.sheetnames:
        if sheet == "Amount of stroke":
            sheet_exists = True
            break

    if sheet_exists:
        ws = wb["Amount of stroke"]
    else:
        ws = wb.create_sheet("Amount of stroke")
    
    # 全シートのB列のデータを収集
    all_data = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Amount of stroke":
            continue  # 自分自身は除外
        
        sheet = wb[sheet_name]
        
        # B列のデータを取得（ヘッダー行を除く）
        for row in range(2, sheet.max_row + 1):
            col_b = sheet[f'B{row}'].value
            
            if col_b is not None:
                all_data.append({
                    'sheet': sheet_name,
                    'col_b': col_b
                })
    
    # データを整理して棒グラフ用のデータを作成
    if all_data:
        # シート名ごとにデータをグループ化
        sheet_summary = {}
        for data in all_data:
            sheet_name = data['sheet']
            if sheet_name not in sheet_summary:
                sheet_summary[sheet_name] = []
            sheet_summary[sheet_name].append(data)
        
        # 棒グラフ用のデータを作成
        graph_data = []
        for sheet_name, data_list in sheet_summary.items():
            # B列の値を合計
            total_b = sum(item['col_b'] for item in data_list if isinstance(item['col_b'], (int, float)))
            
            graph_data.append({
                'sheet_name': sheet_name,
                'total_amount_of_stroke': total_b
            })
        
        # 棒グラフを作成
        if graph_data:
            # ヘッダーを追加
            ws['A1'] = 'Sheet Name'
            ws['B1'] = 'Total Amount of stroke'
            
            # データを書き込み
            for i, data in enumerate(graph_data, start=2):
                ws[f'A{i}'] = data['sheet_name']
                ws[f'B{i}'] = data['total_amount_of_stroke']
            
            # 棒グラフを作成
            from openpyxl.chart import BarChart, Reference
            
            chart = BarChart()
            chart.title = "Comparison of B Column Values Across Sheets"
            chart.x_axis.title = "Sheet Names"
            chart.y_axis.title = "Total Amount of stroke"
            
            # データ範囲を指定
            data = Reference(ws, min_col=2, min_row=1, max_row=len(graph_data)+1, max_col=2)
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(graph_data)+1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # グラフを配置
            ws.add_chart(chart, "D2")
            
            print(f"棒グラフを作成しました: {len(graph_data)}シートのデータを比較")
        else:
            print("グラフ用のデータが見つかりませんでした")
    else:
        print("シートからデータを取得できませんでした")
    
    # ファイルを保存
    wb.save(file_path)

def create_line_graph(folder_path:str) -> List[str]:
    file_path = Path(f"{folder_path}/stroke_count.xlsx")
    wb = openpyxl.load_workbook(file_path)

    try:
        ws = wb["Amount of stroke"]
    except Exception as e:
        raise KeyError(f"シートが見つかりませんでした: {e}")

    # 全シートのA列（日にち）と最後の列（打鍵数）の最新データを収集
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    all_data = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        all_data.append(row)
    
    
    # データを整理して折れ線グラフ用のデータを作成
    if all_data:

        max_col = ws.max_column
        
         # ファイルを保存
        wb.save(file_path)
        import xlsxwriter
        # 新しいExcelファイルを作成

        workbook = xlsxwriter.Workbook(f"{folder_path}/Stroke_Count_Line_Chart.xlsx")
        sheet_name = "stroke_count"
        worksheet = workbook.add_worksheet(sheet_name)

        for i, row in enumerate(all_data):
            if i == 0:
                worksheet.write_row(0, 0, header_row)  # B1:C1
            else:
                worksheet.write_row(i, 0, row)  # B1:C1

        # 折れ線グラフ
        chart = workbook.add_chart({"type": "line"})

        # X軸カテゴリはヘッダ行（B1:C1）
        cat = f"={sheet_name}!$B$1:{get_column_letter(max_col)}$1"

        # 各行を1系列として追加（名前はA列、値はB〜C列）
        for i in range(len(all_data)-1):
            row = i + 2  # Excelは1始まり、データ開始は2行目
            
            chart.add_series({
                "name":       f"={sheet_name}!$A${row}",
                "categories": cat,
                "values":     f"={sheet_name}!$B${row}:${get_column_letter(max_col)}${row}",
                "marker":     {"type": "automatic"},  # 見やすさ向上用（任意）
            })

        # 体裁
        chart.set_title({"name": "Stroke Count"})
        chart.set_legend({"position": "bottom"})
        chart.set_x_axis({"name": "Date"})
        chart.set_y_axis({"name": "Stroke Count"})
        chart.set_size({"width": 720, "height": 420})

        # グラフ挿入
        worksheet.insert_chart("G20", chart)
        workbook.close()



        print("折れ線グラフを作成しました")
        return [f"{folder_path}/Stroke_Count_Line_Chart.xlsx"]
    else:
        print("シートからデータを取得できませんでした")
    


