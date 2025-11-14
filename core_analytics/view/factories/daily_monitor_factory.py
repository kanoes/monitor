"""
Factory for creating daily monitor reports using Excel templates.
"""
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
import logging
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
import pytz
from openpyxl.utils import column_index_from_string

from core_analytics.core.models import ProcessData
from core_analytics.view.excel_utils import _convert_timezone_aware_datetimes
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

class DailyMonitorFactory:
    """Factory for creating daily monitor reports using Excel templates."""
    
    def __init__(self):
        self.logger = logging.getLogger("CoreAnalytics")
        self.template_dir = Path("./config/report_template")
    
    def _copy_template_to_output(self, template_name: str, output_dir: Path, end_time: datetime, output_suffix: str = "report") -> Path:
        """Copy Excel template to output directory with date."""
        template_path = self.template_dir / template_name
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        output_filename = f"daily_monitor_{output_suffix}_{end_time.strftime('%Y%m%d')}.xlsx"
        output_path = output_dir / output_filename
        
        shutil.copy2(template_path, output_path)
        self.logger.info(f"Copied template {template_name} to {output_path}")
        
        return output_path
    
    def _fill_template_with_data(
        self,
        template_path: Path,
        processed_data: ProcessData,
        target_date: datetime,
        mtd_costs: Optional[Dict[str, float]] = None,
    ) -> None:
        """Fill Excel template with query results for a specific date."""
        try:
            workbook = load_workbook(template_path)

            jst = pytz.timezone('Asia/Tokyo')
            current_date = target_date.astimezone(jst)

            self._ensure_month_sheet_exists(workbook, current_date)

            self._fill_daily_data(workbook, processed_data, current_date, mtd_costs)

            workbook.save(template_path)
            self.logger.info(f"Successfully filled template: {template_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to fill template {template_path}: {e}")
            raise

    def _ensure_month_sheet_exists(self, workbook: openpyxl.Workbook, current_date: datetime) -> None:
        """Ensure current month sheet exists, copying template or previous month when needed."""
        try:
            from copy import copy
            from datetime import timedelta
            import calendar

            sheet_name = f"{current_date.year}年{current_date.month}月"
            if sheet_name in workbook.sheetnames:
                return

            template_sheet_name = "YYYY年MM月"
            source_ws = None

            if template_sheet_name in workbook.sheetnames:
                source_ws = workbook[template_sheet_name]
            else:
                prev_month_last_day = (current_date.replace(day=1) - timedelta(days=1))
                prev_name = f"{prev_month_last_day.year}年{prev_month_last_day.month}月"
                if prev_name in workbook.sheetnames:
                    source_ws = workbook[prev_name]

            if source_ws:
                new_ws = workbook.copy_worksheet(source_ws)
            else:
                new_ws = workbook.create_sheet()

            new_ws.title = sheet_name

            sheets = workbook._sheets
            idx_new = sheets.index(new_ws)
            sheets.insert(0, sheets.pop(idx_new))

            ranges = [("A", "E"), ("G", "M"), ("O", "S"), ("U", "AA"), ("AC", "AE")]
            for start_col, end_col in ranges:
                start_idx = column_index_from_string(start_col)
                end_idx = column_index_from_string(end_col)
                for r in range(4, 41):
                    for c in range(start_idx, end_idx + 1):
                        cell = new_ws.cell(row=r, column=c)

                        saved_number_format = copy(cell.number_format)
                        saved_border = copy(cell.border)
                        saved_fill = copy(cell.fill)
                        saved_font = copy(cell.font)
                        saved_alignment = copy(cell.alignment)
                        saved_protection = copy(cell.protection)

                        cell.value = None
                        if cell.hyperlink:
                            cell.hyperlink = None

                        cell.number_format = saved_number_format
                        cell.border = saved_border
                        cell.fill = saved_fill
                        cell.font = saved_font
                        cell.alignment = saved_alignment
                        cell.protection = saved_protection

            last_day_num = calendar.monthrange(current_date.year, current_date.month)[1]

            row = 4
            for d in range(1, last_day_num + 1):
                if row > 40:
                    break
                label = f"{current_date.month}月{d}日"
                for col in ("A", "G", "O", "U", "AC"):
                    cell = new_ws[f"{col}{row}"]
                    saved_number_format = copy(cell.number_format)
                    saved_border = copy(cell.border)
                    saved_fill = copy(cell.fill)
                    saved_font = copy(cell.font)
                    saved_alignment = copy(cell.alignment)
                    
                    cell.value = label
                    
                    cell.number_format = saved_number_format
                    cell.border = saved_border
                    cell.fill = saved_fill
                    cell.font = saved_font
                    cell.alignment = saved_alignment
                    
                row += 1
                
            self.logger.info(f"Created new month sheet '{sheet_name}' with all formats preserved")
            
        except Exception as e:
            self.logger.error(f"Failed to ensure month sheet exists: {e}")
    
    def _fill_daily_data(self, workbook: openpyxl.Workbook, processed_data: ProcessData, current_date: datetime, mtd_costs: Optional[Dict[str, float]] = None) -> None:
        """Fill daily data into the appropriate sheet and cells based on current date."""
        try:
            from copy import copy
            
            def set_cell_value_keep_format(cell, value):
                """Set cell value while keeping all formats."""
                saved_number_format = copy(cell.number_format)
                saved_border = copy(cell.border)
                saved_fill = copy(cell.fill)
                saved_font = copy(cell.font)
                saved_alignment = copy(cell.alignment)
                saved_protection = copy(cell.protection)
                
                cell.value = value
                
                cell.number_format = saved_number_format
                cell.border = saved_border
                cell.fill = saved_fill
                cell.font = saved_font
                cell.alignment = saved_alignment
                cell.protection = saved_protection
            
            sheet_name = f"{current_date.year}年{current_date.month}月"
            
            if sheet_name not in workbook.sheetnames:
                self.logger.warning(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")
                return
            
            ws = workbook[sheet_name]
            
            target_row = self._find_date_row(ws, current_date)
            if target_row is None:
                self.logger.warning(f"Could not find row for date {current_date.strftime('%m月%d日')} in sheet {sheet_name}")
                return
            
            column_mappings = {
                "daily_alm_chat_count": ("B", "C"),
                "daily_alm_dashboard_count": ("D", "E"),
                "daily_doc_search_count": ("H", "I"),
                "daily_my_assistant_search_count": ("P", "Q"),
                "daily_my_assistant_upload_count": ("R", "S"),
                "daily_market_report_web_count": ("V", "W"),
                "daily_company_analyze_count": ("X", "Y"),
                "daily_market_report_bot_count": ("Z", "AA"),
                "daily_brain_count": ("AD", "AE"),
            }
            
            for query_key, (x_col, y_col) in column_mappings.items():
                try:
                    result = self._get_query_result(processed_data, query_key)
                    
                    if result and result.get("data") and result["data"].tables:
                        rows = result["data"].tables[0].rows
                        if rows and len(rows) > 0 and len(rows[0]) >= 2:
                            x_value = rows[0][0]
                            y_value = rows[0][1]
                            
                            set_cell_value_keep_format(ws[f"{x_col}{target_row}"], x_value)
                            set_cell_value_keep_format(ws[f"{y_col}{target_row}"], y_value)
                            
                            self.logger.info(f"Filled {query_key}: [{x_value}, {y_value}] -> {x_col}{target_row}, {y_col}{target_row}")
                        else:
                            set_cell_value_keep_format(ws[f"{x_col}{target_row}"], 0)
                            set_cell_value_keep_format(ws[f"{y_col}{target_row}"], 0)
                            self.logger.warning(f"No data for {query_key}, filled with [0, 0]")
                    else:
                        set_cell_value_keep_format(ws[f"{x_col}{target_row}"], 0)
                        set_cell_value_keep_format(ws[f"{y_col}{target_row}"], 0)
                        self.logger.warning(f"No result for {query_key}, filled with [0, 0]")
                        
                except Exception as e:
                    self.logger.error(f"Failed to fill data for {query_key}: {e}")
                    ws[f"{x_col}{target_row}"] = "ERROR"
                    ws[f"{y_col}{target_row}"] = "ERROR"
            
            if mtd_costs:
                try:
                    cost_items = list(mtd_costs.items())
                    cost_cols = ["J", "K", "L", "M"]

                    for idx, (name, val) in enumerate(cost_items[:len(cost_cols)]):
                        set_cell_value_keep_format(ws[f"{cost_cols[idx]}{target_row}"], val)

                    for col in cost_cols[len(cost_items):]:
                        set_cell_value_keep_format(ws[f"{col}{target_row}"], 0)

                    try:
                        msg_pairs = []
                        for i, (n, v) in enumerate(cost_items[:len(cost_cols)]):
                            msg_pairs.append(f"{n}={v} -> {cost_cols[i]}{target_row}")
                        if msg_pairs:
                            self.logger.info("Filled MTD costs: " + ", ".join(msg_pairs))
                    except Exception:
                        pass

                except Exception as e:
                    self.logger.error(f"Failed to fill MTD costs: {e}")
                    for col in ["J", "K", "L", "M"]:
                        ws[f"{col}{target_row}"] = "ERROR"
            
        except Exception as e:
            self.logger.error(f"Failed to fill daily data: {e}")
    
    def _find_date_row(self, worksheet, current_date: datetime) -> int:
        """Find the row number that matches the current date in column A."""
        try:
            target_date_str = f"{current_date.month}月{current_date.day}日"
            
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet[f"A{row}"].value
                if cell_value and str(cell_value).strip() == target_date_str:
                    return row
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error finding date row: {e}")
            return None
    
    def _get_query_result(self, processed_data: ProcessData, query_key: str):
        """Get query result from processed data."""
        for results_dict in [processed_data.user_count_results, 
                           processed_data.stroke_count_results, 
                           processed_data.unknown_results]:
            if query_key in results_dict:
                return results_dict[query_key]
        return None
    
    def _fill_history_template_with_data(self, template_path: Path, processed_data: ProcessData) -> None:
        """Fill history Excel template with query results."""
        try:
            workbook = load_workbook(template_path)
            
            self._fill_history_data_to_sheets(workbook, processed_data)
            
            workbook.save(template_path)
            self.logger.info(f"Successfully filled history template: {template_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to fill history template {template_path}: {e}")
            raise
    
    def _fill_history_data_to_sheets(self, workbook: openpyxl.Workbook, processed_data: ProcessData) -> None:
        """Fill history data into specific sheets."""
        try:
            sheet_mappings = {
                "daily_alm_chat_history": "外貨ALM-Chat",
                "daily_alm_dashboard_history": "外貨ALM-Dashboard", 
                "daily_doc_search_history": "ドキュメントサーチ",
                "daily_my_assistant_search_history": "My Assistant-検索",
                "daily_my_assistant_upload_history": "My Assistant-アップロード",
                "daily_market_report_web_history": "マーケット分析Web",
                "daily_market_report_bot_history": "マーケット分析Bot",
                "daily_company_analyze_history": "会社情報",
                "daily_brain_history": "為替分析Brain",
            }
            
            def _sanitize_excel_value(value, max_len: int = 32767):
                if value is None:
                    return None
                from datetime import datetime as _dt, date as _date, time as _time
                if isinstance(value, (_dt, _date, _time)):
                    return value
                if isinstance(value, (int, float)):
                    return value
                if isinstance(value, (bytes, bytearray)):
                    try:
                        value = value.decode("utf-8", errors="ignore")
                    except Exception:
                        value = str(value)
                if not isinstance(value, str):
                    value = str(value)
                value = ILLEGAL_CHARACTERS_RE.sub("", value)
                if len(value) > max_len:
                    value = value[:max_len]
                return value

            for query_key, sheet_name in sheet_mappings.items():
                try:
                    if sheet_name not in workbook.sheetnames:
                        self.logger.warning(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")
                        continue
                    
                    ws = workbook[sheet_name]
                    result = self._get_query_result(processed_data, query_key)
                    
                    if result and result.get("data") and result["data"].tables:
                        self._clear_worksheet_data(ws)
                        
                        df = pd.DataFrame(
                            data=result["data"].tables[0].rows,
                            columns=result["data"].tables[0].columns
                        )
                        df = _convert_timezone_aware_datetimes(df)
                        
                        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
                            for col_idx, value in enumerate(row_data, start=1):
                                ws.cell(row=row_idx, column=col_idx, value=_sanitize_excel_value(value))
                        
                        self.logger.info(f"Filled {len(df)} rows for {query_key} in sheet '{sheet_name}'")
                    else:
                        self.logger.warning(f"No data found for {query_key}")
                        
                except Exception as e:
                    self.logger.error(f"Failed to fill data for {query_key} in sheet '{sheet_name}': {e}")
                    continue
            
        except Exception as e:
            self.logger.error(f"Failed to fill history data to sheets: {e}")
    
    def _clear_worksheet_data(self, worksheet) -> None:
        """Clear worksheet data while keeping the first row as headers."""
        try:
            max_row = worksheet.max_row
            if max_row > 1:
                for row in worksheet.iter_rows(min_row=2, max_row=max_row):
                    for cell in row:
                        cell.value = None
                        if cell.hyperlink:
                            cell.hyperlink = None
        except Exception as e:
            self.logger.error(f"Failed to clear worksheet data: {e}")
    
    def generate_daily_monitor_report(self, processed_data: ProcessData, output_dir: str, end_time: datetime, mtd_costs: Optional[Dict[str, float]] = None) -> List[str]:
        """Generate daily monitor report using Excel templates."""
        self.logger.info("Starting daily monitor report generation")
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        generated_files = []
        
        try:
            usage_report_path = self._generate_cumulative_usage_report(output_path, processed_data, end_time, mtd_costs)
            generated_files.append(str(usage_report_path))
            self.logger.info(f"Successfully generated cumulative usage report: {usage_report_path}")
            
            history_report_path = self._generate_daily_history_report(output_path, processed_data, end_time)
            generated_files.append(str(history_report_path))
            self.logger.info(f"Successfully generated daily history report: {history_report_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to generate daily monitor reports: {e}")
            raise
        
        return generated_files
    
    def _generate_cumulative_usage_report(self, output_path: Path, processed_data: ProcessData, end_time: datetime, mtd_costs: Optional[Dict[str, float]] = None) -> Path:
        """Generate cumulative usage report (updates existing file)."""
        base_dir = Path("output/市場GAI打鍵")
        base_dir.mkdir(parents=True, exist_ok=True)
        usage_report_path = base_dir / "市場GAI使用状況.xlsx"
        
        if usage_report_path.exists():
            self.logger.info(f"Found existing usage report, updating: {usage_report_path}")
        else:
            self.logger.info(f"Creating new usage report from template: {usage_report_path}")
            template_path = self.template_dir / "市場GAI使用状況.xlsx"
            if not template_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_path}")
            shutil.copy2(template_path, usage_report_path)
        
        self._fill_template_with_data(usage_report_path, processed_data, end_time, mtd_costs)

        return usage_report_path
    
    def _generate_daily_history_report(self, output_path: Path, processed_data: ProcessData, end_time: datetime) -> Path:
        """Generate daily history report (creates new file each day)."""
        jst = pytz.timezone('Asia/Tokyo')
        current_date = datetime.now(jst)
        
        date_str = current_date.strftime('%Y%m%d')
        ym_str = current_date.strftime('%Y%m')

        base_dir = Path("output/市場GAI打鍵/打鍵詳細履歴") / ym_str
        base_dir.mkdir(parents=True, exist_ok=True)
        history_report_path = base_dir / f"市場GAI打鍵履歴_{date_str}.xlsx"
        
        template_path = self.template_dir / "市場GAI打鍵履歴_YYYYMMDD.xlsx"
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        shutil.copy2(template_path, history_report_path)
        self.logger.info(f"Created new history report from template: {history_report_path}")
        
        self._fill_history_template_with_data(history_report_path, processed_data)
        
        return history_report_path
