"""
Test script for rebuild command.
This script tests the rebuild functionality with:
- Real Azure Log Analytics data fetching
- Mocked email service (no emails sent)
- Mocked blob upload (files copied to test/result instead)
"""
import os
import sys
import shutil
from pathlib import Path
from unittest.mock import Mock, patch
from datetime import datetime
from zoneinfo import ZoneInfo

# Add project root to path
project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root))
os.chdir(project_root)

from core_analytics.config.settings import ConfigurationService
from core_analytics.model.repositories.azure_log_repository import AzureLogRepository
from core_analytics.services.analytics_service import AnalyticsService
from core_analytics.services.query_strategies.strategy_factory import QueryStrategyFactory
from core_analytics.view.factories.daily_monitor_factory import DailyMonitorFactory
from core_analytics.core.logging_config import LoggerSetup
from core_analytics.command.rebuild import (
    parse_date,
    generate_date_range,
    initialize_usage_report_from_template,
    process_single_day
)

logger = LoggerSetup.setup_logger()


class MockBlobRepository:
    """Mock blob repository that copies files to test/result instead of uploading."""
    
    def __init__(self, config_service: ConfigurationService):
        self.config_service = config_service
        self.logger = logger
        self.result_dir = project_root / "test" / "result"
        self.result_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"MockBlobRepository initialized. Files will be copied to {self.result_dir}")
    
    def upload_file(self, local_file_path: str, remote_path: str) -> str:
        """Copy file to test/result instead of uploading to blob."""
        try:
            local_path = Path(local_file_path)
            if not local_path.exists():
                self.logger.error(f"Local file not found: {local_file_path}")
                return ""
            
            # Preserve directory structure in result folder
            # remote_path might be like: "output/市場GAI打鍵/打鍵詳細履歴/202411/市場GAI打鍵履歴_20241113.xlsx"
            # or just the relative path from project root
            relative_path = Path(remote_path)
            
            # If path contains "output", extract everything after "output"
            if "output" in str(relative_path):
                parts = relative_path.parts
                if "output" in parts:
                    idx = parts.index("output")
                    relative_path = Path(*parts[idx:])  # Keep "output" and everything after
                else:
                    # Find "output" in the string path
                    path_str = str(relative_path)
                    if "output" in path_str:
                        output_idx = path_str.find("output")
                        relative_path = Path(path_str[output_idx:])
                    else:
                        relative_path = relative_path.name
            elif relative_path.is_absolute():
                # If absolute path, try to extract relative part
                project_root_str = str(project_root)
                if project_root_str in str(relative_path):
                    relative_path = Path(str(relative_path).replace(project_root_str, "").lstrip("/"))
                else:
                    relative_path = relative_path.name
            
            dest_path = self.result_dir / relative_path
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            
            shutil.copy2(local_path, dest_path)
            self.logger.info(f"Copied {local_file_path} to {dest_path} (mock upload)")
            return str(dest_path)
            
        except Exception as e:
            self.logger.error(f"Failed to copy file {local_file_path}: {e}")
            return ""


class MockEmailService:
    """Mock email service that doesn't send emails."""
    
    def __init__(self):
        self.logger = logger
        logger.info("MockEmailService initialized. No emails will be sent.")
    
    def send_daily_monitor_report(self, file_paths: list, date_str: str) -> bool:
        """Mock email sending - just log the action."""
        self.logger.info(f"Mock: Would send email with {len(file_paths)} files for date range: {date_str}")
        for file_path in file_paths:
            self.logger.info(f"  - {file_path}")
        return True


def test_rebuild(rebuild_from: str, rebuild_to: str):
    """
    Test rebuild functionality with mocked email and blob services.
    
    Args:
        rebuild_from: Start date in YYYYMMDD format
        rebuild_to: End date in YYYYMMDD format
    """
    logger.info(f"Starting test rebuild process from {rebuild_from} to {rebuild_to}")
    
    # Validate date format
    try:
        parse_date(rebuild_from)
        parse_date(rebuild_to)
    except ValueError as e:
        logger.error(f"Invalid date format: {e}")
        raise
    
    date_range = generate_date_range(rebuild_from, rebuild_to)
    logger.info(f"Processing {len(date_range)} days")
    
    # Initialize services with real Azure Log Analytics
    config_service = ConfigurationService(days_range=1)
    
    # Override output_base_dir to use test/result
    original_output_dir = config_service.get_app_settings().output_base_dir
    test_output_dir = project_root / "test" / "result" / "output"
    test_output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create a modified AppSettings with test output directory
    from core_analytics.config.settings import AppSettings
    app_settings = AppSettings(query_days_range=1)
    app_settings.output_base_dir = str(test_output_dir)
    
    # Patch the config service to return our modified settings
    config_service._app_settings = app_settings
    
    # Initialize real Azure services
    log_repository = AzureLogRepository(config_service)
    strategy_factory = QueryStrategyFactory()
    analytics_service = AnalyticsService(
        log_repository,
        config_service,
        strategy_factory
    )
    daily_monitor_factory = DailyMonitorFactory()
    
    # Use mock blob repository
    storage_service = MockBlobRepository(config_service)
    
    # Use mock email service
    email_service = MockEmailService()
    
    all_generated_files = []
    
    # Initialize usage report
    usage_report_path = test_output_dir / "市場GAI打鍵" / "市場GAI使用状況.xlsx"
    usage_report_path.parent.mkdir(parents=True, exist_ok=True)
    
    template_path = project_root / "config" / "report_template" / "市場GAI使用状況.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    initialize_usage_report_from_template(template_path, usage_report_path)
    
    # Process each day
    for target_date in date_range:
        try:
            logger.info(f"\n{'='*60}")
            logger.info(f"Processing date: {target_date.strftime('%Y-%m-%d')}")
            logger.info(f"{'='*60}")
            
            # Check if processed_data has data before processing
            logger.info(f"Fetching data for {target_date.strftime('%Y-%m-%d')}...")
            
            # Pre-check: verify we can fetch data
            jst = ZoneInfo("Asia/Tokyo")
            target_date_jst = target_date.astimezone(jst)
            start_time = target_date_jst.replace(hour=0, minute=0, second=0, microsecond=0)
            end_time = target_date_jst.replace(hour=23, minute=59, second=59, microsecond=999999)
            start_time_utc = start_time.astimezone(datetime.UTC)
            end_time_utc = end_time.astimezone(datetime.UTC)
            
            try:
                test_data = analytics_service.fetch_and_process_data(start_time_utc, end_time_utc)
                logger.info(f"Fetched data: user_count_results={len(test_data.user_count_results)}, "
                          f"stroke_count_results={len(test_data.stroke_count_results)}, "
                          f"unknown_results={len(test_data.unknown_results)}")
            except Exception as e:
                logger.error(f"Failed to fetch data: {e}", exc_info=True)
            
            files = process_single_day(
                target_date,
                config_service,
                log_repository,
                strategy_factory,
                analytics_service,
                daily_monitor_factory,
                storage_service
            )
            
            logger.info(f"process_single_day returned {len(files)} files: {files}")
            
            # Verify files were actually created and have content
            for file_path in files:
                file_path_obj = Path(file_path)
                if file_path_obj.exists():
                    file_size = file_path_obj.stat().st_size
                    logger.info(f"✓ File exists: {file_path_obj} (size: {file_size} bytes)")
                    if file_size < 1000:  # Very small file might be empty
                        logger.warning(f"⚠ File is very small, might be empty: {file_path_obj}")
                else:
                    logger.error(f"✗ File NOT found: {file_path_obj}")
            
            history_files = [f for f in files if "打鍵履歴" in f]
            all_generated_files.extend(history_files)
            
            # Copy generated files to test/result (in case they were created in output directory)
            for file_path in files:
                file_path_obj = Path(file_path)
                if file_path_obj.exists():
                    logger.info(f"Found generated file: {file_path_obj} (size: {file_path_obj.stat().st_size} bytes)")
                    # If file is in output directory, copy it to test/result preserving structure
                    if "output" in str(file_path_obj):
                        storage_service.upload_file(str(file_path_obj), str(file_path_obj))
                else:
                    logger.warning(f"File not found: {file_path_obj}")
            
            logger.info(f"Successfully processed {target_date.strftime('%Y-%m-%d')}")
            
        except Exception as e:
            logger.error(f"Failed to process date {target_date.strftime('%Y%m%d')}: {e}", exc_info=True)
            continue
    
    # Copy final usage report to result directory
    # Note: usage_report_path might be in test/result/output already, but also check output directory
    output_usage_path = project_root / "output" / "市場GAI打鍵" / "市場GAI使用状況.xlsx"
    if output_usage_path.exists():
        file_size = output_usage_path.stat().st_size
        logger.info(f"Found usage report in output: {output_usage_path} (size: {file_size} bytes)")
        all_generated_files.append(str(output_usage_path))
        storage_service.upload_file(str(output_usage_path), str(output_usage_path))
        logger.info(f"Final usage report copied to result directory")
        
        # Verify the file has content by checking if it has sheets
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_usage_path)
            logger.info(f"Usage report has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"  Sheet '{sheet_name}': {ws.max_row} rows, {ws.max_column} columns")
        except Exception as e:
            logger.warning(f"Could not inspect usage report: {e}")
    elif usage_report_path.exists():
        file_size = usage_report_path.stat().st_size
        logger.info(f"Found usage report in test output: {usage_report_path} (size: {file_size} bytes)")
        all_generated_files.append(str(usage_report_path))
        storage_service.upload_file(str(usage_report_path), str(usage_report_path))
        logger.info(f"Final usage report copied to result directory")
    else:
        logger.warning(f"Usage report not found in either location!")
        logger.warning(f"  Checked: {output_usage_path}")
        logger.warning(f"  Checked: {usage_report_path}")
    
    # Mock email sending
    if all_generated_files:
        date_str = f"{rebuild_from} to {rebuild_to}"
        email_sent = email_service.send_daily_monitor_report(all_generated_files, date_str)
        if email_sent:
            logger.info("Mock email service logged successfully")
    
    # Final step: Copy any remaining files from output directory to test/result
    output_dir = project_root / "output"
    if output_dir.exists():
        logger.info(f"Copying remaining files from {output_dir} to {storage_service.result_dir}")
        for file_path in output_dir.rglob("*.xlsx"):
            if file_path.is_file():
                file_size = file_path.stat().st_size
                logger.info(f"Found Excel file: {file_path} (size: {file_size} bytes)")
                relative_path = file_path.relative_to(project_root)
                storage_service.upload_file(str(file_path), str(relative_path))
                if str(file_path) not in all_generated_files:
                    all_generated_files.append(str(file_path))
        
        # Also check for usage report specifically
        usage_report_in_output = project_root / "output" / "市場GAI打鍵" / "市場GAI使用状況.xlsx"
        if usage_report_in_output.exists():
            file_size = usage_report_in_output.stat().st_size
            logger.info(f"Found usage report in output: {usage_report_in_output} (size: {file_size} bytes)")
            if str(usage_report_in_output) not in all_generated_files:
                all_generated_files.append(str(usage_report_in_output))
                storage_service.upload_file(str(usage_report_in_output), str(usage_report_in_output.relative_to(project_root)))
    
    logger.info(f"\n{'='*60}")
    logger.info(f"Test rebuild completed!")
    logger.info(f"Generated {len(all_generated_files)} report files")
    logger.info(f"Processed dates from {rebuild_from} to {rebuild_to}")
    logger.info(f"All files saved to: {storage_service.result_dir}")
    logger.info(f"{'='*60}\n")
    
    print(f"✅ Test rebuild completed successfully!")
    print(f"📊 Generated {len(all_generated_files)} report files")
    print(f"📅 Processed dates from {rebuild_from} to {rebuild_to}")
    print(f"📁 Files saved to: {storage_service.result_dir}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Test rebuild command with mocked email and blob services")
    parser.add_argument(
        "--from",
        dest="rebuild_from",
        required=True,
        help="Start date in YYYYMMDD format (e.g., 20241101)"
    )
    parser.add_argument(
        "--to",
        dest="rebuild_to",
        required=True,
        help="End date in YYYYMMDD format (e.g., 20241113)"
    )
    
    args = parser.parse_args()
    
    try:
        test_rebuild(args.rebuild_from, args.rebuild_to)
    except Exception as e:
        logger.error(f"Test rebuild failed: {e}", exc_info=True)
        sys.exit(1)

